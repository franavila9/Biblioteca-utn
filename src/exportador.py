from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def color_argb(color_hex):
    if color_hex.startswith("#"):
        color_hex = color_hex[1:]

    if len(color_hex) == 6:
        return f"FF{color_hex}"

    return color_hex


def crear_fill(color_hex):
    return PatternFill(
        fill_type="solid",
        fgColor=color_argb(color_hex)
    )


def exportar_a_excel(
    df_mostrado,
    df_completo,
    resultados,
    ruta_salida="output/vector_estado.xlsx"
):
    ruta = Path(ruta_salida)
    ruta.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    ws_mostrado = wb.active
    ws_mostrado.title = "Vector mostrado"
    escribir_vector_con_formato(ws_mostrado, df_mostrado)

    ws_completo = wb.create_sheet("Vector completo")
    escribir_vector_con_formato(ws_completo, df_completo)

    ws_resultados = wb.create_sheet("Resultados")
    escribir_resultados(ws_resultados, resultados)

    wb.save(ruta)

    return ruta


def obtener_bloques(columnas):
    bloques = []

    def agregar(nombre_general, nombre_sub, columnas_bloque, color):
        columnas_existentes = [col for col in columnas_bloque if col in columnas]

        if columnas_existentes:
            bloques.append({
                "general": nombre_general,
                "sub": nombre_sub,
                "columnas": columnas_existentes,
                "color": color
            })

    agregar(
        "Identificación",
        "",
        ["fila", "evento", "reloj"],
        "EDEDED"
    )

    agregar(
        "Eventos",
        "Llegada_persona",
        [
            "rnd_tipo_tramite",
            "tipo_tramite",
            "tiempo_entre_llegadas",
            "proxima_llegada",
        ],
        "D9EAF7"
    )

    agregar(
        "Eventos",
        "Fin_atención(i)",
        [
            "rnd_atencion",
            "tiempo_atencion",
            "fin_atencion(1)",
            "fin_atencion(2)",
        ],
        "FFF2CC"
    )

    agregar(
        "Eventos",
        "fin_lectura",
        [
            "rnd_post_prestamo",
            "post_prestamo",
            "rnd_tiempo_lectura",
            "tiempo_lectura",
            "proximo_fin_lectura",
        ],
        "C6EFCE"
    )

    agregar(
        "Objetos permanentes",
        "Empleado(N)",
        [
            "empleado(1)_estado",
            "empleado(2)_estado",
            "cola",
        ],
        "DDEBF7"
    )

    agregar(
        "Objetos permanentes",
        "Biblioteca",
        [
            "cantidad_personas",
            "estado_biblioteca",
        ],
        "DDEBF7"
    )

    agregar(
        "Variables estadísticas",
        "Estadísticas",
        [
            "acum_tiempo_permanencia",
            "contador_clientes_salieron",
            "tiempo_promedio_permanencia",
            "ac_ocio_empleado_1",
            "ac_ocio_empleado_2",
        ],
        "FCE4D6"
    )

    columnas_clientes = [col for col in columnas if col.startswith("cli(")]

    agregar(
        "Objetos temporales",
        "Clientes",
        columnas_clientes,
        "D9EAD3"
    )

    return bloques


def escribir_vector_con_formato(ws, df):
    columnas = list(df.columns)
    bloques = obtener_bloques(columnas)

    thin = Side(border_style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_index = 1

    for bloque in bloques:
        inicio = col_index
        fin = col_index + len(bloque["columnas"]) - 1

        fill = crear_fill(bloque["color"])

        if inicio != fin:
            ws.merge_cells(
                start_row=1,
                start_column=inicio,
                end_row=1,
                end_column=fin
            )

        celda_general = ws.cell(row=1, column=inicio)
        celda_general.value = bloque["general"]
        celda_general.fill = fill
        celda_general.font = header_font
        celda_general.alignment = center
        celda_general.border = border

        if inicio != fin:
            ws.merge_cells(
                start_row=2,
                start_column=inicio,
                end_row=2,
                end_column=fin
            )

        celda_sub = ws.cell(row=2, column=inicio)
        celda_sub.value = bloque["sub"]
        celda_sub.fill = fill
        celda_sub.font = header_font
        celda_sub.alignment = center
        celda_sub.border = border

        for columna in range(inicio, fin + 1):
            ws.cell(row=1, column=columna).fill = fill
            ws.cell(row=1, column=columna).border = border
            ws.cell(row=2, column=columna).fill = fill
            ws.cell(row=2, column=columna).border = border

        for columna_nombre in bloque["columnas"]:
            celda = ws.cell(row=3, column=col_index)
            celda.value = columna_nombre
            celda.fill = fill
            celda.font = header_font
            celda.alignment = center
            celda.border = border
            col_index += 1

    for row_idx, (_, fila) in enumerate(df.iterrows(), start=4):
        for col_idx, columna in enumerate(columnas, start=1):
            celda = ws.cell(row=row_idx, column=col_idx)
            celda.value = fila[columna]
            celda.alignment = center
            celda.border = border

    ws.freeze_panes = "A4"

    for col_idx, columna in enumerate(columnas, start=1):
        max_len = len(str(columna))

        for row in range(4, min(ws.max_row, 100) + 1):
            value = ws.cell(row=row, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))

        if columna == "evento":
            ancho = 32
        elif columna.startswith("cli("):
            ancho = 18
        else:
            ancho = min(max_len + 2, 28)

        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 35


def escribir_resultados(ws, resultados):
    ws["A1"] = "Resultado"
    ws["B1"] = "Valor"

    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    fila = 2

    for clave, valor in resultados.items():
        ws.cell(row=fila, column=1).value = clave
        ws.cell(row=fila, column=2).value = valor
        fila += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20